import re
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Añadimos los contactos de email
email_TO = ';santos-sanchez@eipsa.es;'
email_TO_CC = ';jesus-martinez@eipsa.es;ernesto-carrillo@eipsa.es;'
email_LB = ';luis-bravo@eipsa.es;'
email_AC = ';ana-calvo@eipsa.es;'
email_SS = ';sandra-sanz@eipsa.es;'
email_JV = ';jorge-valtierra@eipsa.es'
email_CC = ';carlos-crespohor@eipsa.es;'


def prodoc_vendor_number(df):
    """
    Función para cambiar el tipo de documento a entero y añadir la hora exacta recibida del email.

    Args:
        df (pandas.DataFrame): DataFrame que contiene las columnas 'Tipo de documento' y 'Fecha'.
        receivedtime (datetime): Hora exacta recibida del email.

    Returns:
        pandas.DataFrame: DataFrame actualizado con el tipo de documento cambiado a entero y la hora exacta añadida.
    """
    # mapping (dict): Diccionario de mapeo para identificar el tipo de documento
    mapping = {'7011318362': 'P-24/091', '7070000087': 'P-24/054',
               '7011319592': 'P-24/073', '7011294464': 'P-23/087',
               }

    # Asegurar que la columna es string y eliminar espacios en blanco
    df['Nº Pedido'] = df['Nº Pedido'].astype(str).str.strip()

    # Aplicar mapeo y mantener valores originales si no están en el diccionario
    df['Nº Pedido'] = df['Nº Pedido'].map(mapping).fillna(df['Nº Pedido'])

    return df


def reemplazar_null(df):
    """
        Esta función toma un DataFrame como entrada y reemplaza los valores de la columna "Suplemento" de acuerdo con el mapeo proporcionado en el diccionario mapping

        Args:
            df (pandas.DataFrame): DataFrame que contiene "NULOS".

        Returns:
            pandas.DataFrame: DataFrame actualizado, si el valor no se encuentra en el mapeo o es NaN, se reemplaza con 'S00'.
    """
    mapping = {np.nan: 'S00', 'S01': 'S01', 'S02': 'S02', 'S03': 'S03',
               'S04': 'S04', 'S05': 'S05', 'S06': 'S06', 'S07': 'S07'}
    df['Supp.'] = df['Supp.'].map(mapping).fillna('S00')
    return df


def identificar_cliente_por_PO(df):
    """
    Función para identificar el cliente a través del número de pedido (PO) utilizando expresiones regulares.

    Args:
        df (pandas.DataFrame): DataFrame que contiene la columna 'PO'.

    Returns:
        pandas.DataFrame: DataFrame actualizado con la columna 'Cliente' identificada.
    """
    # mapping (dict): Diccionario de mapeo para identificar el cliente según el número de pedido.
    mapping = {'10121': 'DUQM', '10150': 'BAPCO',
               '10160': 'CRISP', '10230': 'MARJAN',
               '10318': 'RAS TANURA', '10330': 'NEW PTA COMPLEX',
               '10370': 'QATAR EPC3', '10380': 'YPF',
               '10400': 'ADNOC DALMA', '10430': 'QATAR EPC4',
               '23222': 'CQP', '23262': 'Certificado',
               '33138': 'DUQM', '70150': 'SEWA',
               '70215': 'CFE MERIDA', '70225': 'C.C. VALLADOLID',
               '70230': 'C.C. GONZALEZ ORTEGA', '70240': 'C.C. SAN LUIS',
               '80057': 'BU HASA', '80091': 'T.R. ENAP',
               '19085': 'CEPSA/T.R.', '30011': 'BP OIL ESPAÑA',
               '75001': 'TECNIMONT', '60001': 'CEPSA WOOD',
               '70112': 'CEPSA SAN ROQUE', '70801': 'CEPSA',
               '15282': 'ASTCOR', 'T.206': 'REPSOL PETRÓLEO',
               'BP-T2': 'CNTCC', 'EP24I': 'ALMARAZ/TRILLO',
               '49000': 'JIGPC/ARAMCO', 'PO 15': 'ASTCOR',
               'Q3710': 'INTECSA INDUSTRIAL', 'RFQ 1': 'BU HASA',
               '70292': 'LECTA', 'APEIS': 'KNPC',
               '***': 'CEPSA/AYESA', '30012': 'BP OIL REFINERIA',
               'EC24T': 'ALMARAZ/TRILLO', '10735': 'SULZER',
               '70700': 'CEPSA/WOOD', 'JUS&I': 'ARAMCO/HYUNDAI',
               '70113': 'CEPSA', '10620': 'QATARBOP/TR',
               'ADI-29': 'TECHNIP/SYNKEDIA', '10431': 'QATAREPC4/TR',
               'PO P7': 'TECHNIP/REPSOL', '12574': 'ALPARGATA',
               'ADI-2': 'TECHNIP/SYNKEDIA', '23000': 'TECHNIP/GALP',
               '45077': 'ARAMCO PORTAL', '45000': 'AYESA/REPSOL',
               '30015': 'BP OIL ESPAÑA', '19162': 'WISON/ARAMCO',
               '48550': 'WISON/ARAMCO', '20175': 'TECHNIP/REPSOL',
               'QR-DD': 'ASTCOR/WOOD', 'RFPP-': 'IDOM/REPSOL',
               '10120': 'TR/DUQM', 'SOCAR': 'SOCAR/EMERSON',
               '41650': 'SOCAR/EMERSON', 'P-P0C': 'SACYR/REPSOL',
               'SEG/B': 'SINOPEC/ARAMCO', 'SEG /': 'SINOPEC/ARAMCO',
               '10651': 'ARAMCO/RIYAS', '45124': 'ADNOC/YOKOGAWA',
               'O-23/': 'SINES/YOKOGAWA', 'O-24/': 'SENER/GATE',
               'GAT22': 'SENER/GATE', '45126': 'ADNOC/YOKOGAWA',
               '70700': 'CEPSA/WOOD'}

    # Definir la expresión regular para extraer los primeros 5 dígitos del número de pedido (PO)
    regex_pattern = r'^(\d{5})'

    # Aplicar la expresión regular para extraer los primeros 5 dígitos del PO y mapear el cliente
    df['Cliente'] = df['P.O.'].apply(lambda x: mapping[re.match(regex_pattern, x).group(1)] if re.match(regex_pattern, x) else '')

    return df


def reconocer_tipo_proyecto(df):
    """
    Reconoce el tipo de proyecto basado en el número de pedido ('PO') y lo asigna a la columna 'Material'.

    Args:
        df (pandas.DataFrame): DataFrame que contiene la columna 'PO'.

    Returns:
        pandas.DataFrame: DataFrame actualizado con la columna 'Material' indicando el tipo de proyecto.
    """
    mapping = {
        '7011318362': 'CAUDAL',
        '7070000087': 'TEMPERATURA',
        '7011319592': 'TEMPERATURA',
        '7011294464': 'PLACAS',
    }

    # Asignamos el tipo de proyecto según el número de pedido
    df['Material'] = df['PO'].map(mapping).fillna(df['PO'])  # Si no se encuentra en el mapeo, deja el código original

    return df


def procesar_documento_y_fecha(df, receivedtime):
    """
    Función para cambiar el tipo de documento a entero y añadir la hora exacta recibida del email.

    Args:
        df (pandas.DataFrame): DataFrame que contiene las columnas 'Tipo de documento' y 'Fecha'.
        receivedtime (datetime): Hora exacta recibida del email.

    Returns:
        pandas.DataFrame: DataFrame actualizado con el tipo de documento cambiado a entero y la hora exacta añadida.
    """
    # mapping (dict): Diccionario de mapeo para identificar el tipo de documento
    mapping = {'PLG': 'Planos', 'DWG': 'Planos',
               'CAL': 'Cálculos', 'ESP': 'Cálculos y Planos',
               'CER': 'Certificado', 'NACE': 'Certificado',
               'DOS': 'Dossier', 'LIS': 'Listado',
               'ITP': 'Procedimientos', 'PRC': 'Procedimientos',
               'MAN': 'Manual', 'VDB': 'Listado',
               'PLN': 'PPI', 'PLD': 'Nameplate',
               'CAT': 'Catalogo', 'DL': 'Listado',
               'SPL': 'Repuestos'}

    # Cambiar el tipo de documento usando el mapeo proporcionado
    df['Tipo de documento'] = df['Tipo de documento'].map(mapping)

    # Convertir la hora exacta recibida del email a formato de fecha y hora
    df['Fecha'] = pd.to_datetime(receivedtime, dayfirst=True)

    return df

def critico(df):
    """
    Función para cambiar el tipo de estado en un DataFrame.

    Args:
        df (pandas.DataFrame): DataFrame que contiene la columna 'Return Status'.

    Returns:
        pandas.DataFrame: DataFrame actualizado con los tipos de estado modificados.
    """

    # mapping (dict): Diccionario de mapeo para identificar el estado del documento
    mapping = {'Planos': 'Sí',
               'Cálculos': 'Sí', 'Cálculos y Planos': 'Sí',
               'Certificado': 'No',
               'Dossier': 'No',
               'Procedimientos': 'No',
               'Manual': 'Sí',
               'PPI': 'Sí', 'Nameplate': 'No',
               'Catalogo': 'Sí', 'Listado': 'Sí',
               'Repuestos': 'No'}

    # Aplicar el mapeo para cambiar el tipo de estado en la columna 'Return Status'
    df['Crítico'] = df['Tipo de documento'].map(mapping)

    return df

def cambiar_tipo_estado(df):
    """
    Función para cambiar el tipo de estado en un DataFrame.

    Args:
        df (pandas.DataFrame): DataFrame que contiene la columna 'Return Status'.

    Returns:
        pandas.DataFrame: DataFrame actualizado con los tipos de estado modificados.
    """

    # mapping (dict): Diccionario de mapeo para identificar el estado del documento
    mapping = {
        'A - REJECTED': 'Rechazado',
        '1 - WITH COMMENTS': 'Com. Mayores',
        '2 - WITHOUT COMMENTS': 'Aprobado',
        '2I - FOR INFORMATION ONLY ': 'Informativo',
        '3 - WITH MINOR COMMENTS': 'Com. Menores',
        # AÑADIR PORTAL PRODOC
    }

    # Aplicar el mapeo para cambiar el tipo de estado en la columna 'Return Status'
    df['S.R. Status'] = df['S.R. Status'].map(mapping)

    return df

def email_employee(df):
    """
        Función para identificar el empleado encargado del documento

        Args:
            df (pandas.DataFrame): DataFrame que contiene el Tipo de documento pasado a la nueva columna df2['EMAIL'].

        Returns:
            pandas.DataFrame: DataFrame actualizado con los tipos de documento indicándonos quien es el responsable del documento
    """

    mapping = {'PLG': '', 'DWG': '', 'CAL': '', 'ESP': '', 'CER': email_JV, 'NACE': '', 'LIS': email_JV, 'ITP': '',
               'PRC': email_JV, 'MAN': email_JV, 'VDB': '', 'PLN': '', 'PLD': '', 'CAT': email_JV, 'DL': '', 'DOS': email_JV, 'SPL': email_JV}

    df['EMAIL'] = df['EMAIL'].map(mapping)
    df = df['EMAIL'].apply(pd.Series)
    return df


# Diccionario de mapeo para la función get_responsable_email()
email_mapping = {'P-21/003': email_LB,
                 'P-22/001': email_LB, 'P-22/002': email_LB, 'P-22/003': email_AC, 'P-22/004': email_AC,
                 'P-22/005': email_AC, 'P-22/006': email_LB, 'P-22/007': email_LB, 'P-22/008': email_AC,
                 'P-22/009': email_LB, 'P-22/010': email_AC, 'P-22/011': email_LB, 'P-22/012': email_AC,
                 'P-22/013': email_LB, 'P-22/014': email_AC, 'P-22/015': email_LB, 'P-22/016': email_LB,
                 'P-22/017': email_AC, 'P-22/018': email_AC, 'P-22/019': email_AC, 'P-22/020': email_LB,
                 'P-22/021': email_AC, 'P-22/022': email_AC, 'P-22/023': email_AC, 'P-22/024': email_AC,
                 'P-22/025': email_LB, 'P-22/026': email_LB, 'P-22/027': email_LB, 'P-22/028': email_AC,
                 'P-22/029': email_LB, 'P-22/030': email_LB, 'P-22/031': email_AC, 'P-22/032': email_AC,
                 'P-22/033': email_LB, 'P-22/034': email_LB, 'P-22/035': email_AC, 'P-22/036': email_AC,
                 'P-22/037': email_LB, 'P-22/038': email_AC, 'P-22/039': email_AC, 'P-22/040': email_LB,
                 'P-22/041': email_LB, 'P-22/042': email_AC, 'P-22/043': email_AC, 'P-22/044': email_AC,
                 'P-22/045': email_AC, 'P-22/046': email_AC, 'P-22/047': email_SS, 'P-22/048': email_LB,
                 'P-22/049': email_LB, 'P-22/050': email_LB, 'P-22/051': email_AC, 'P-22/052': email_AC,
                 'P-22/053': email_SS, 'P-22/054': email_SS, 'P-22/055': email_AC, 'P-22/056': email_AC,
                 'P-22/057': email_AC, 'P-22/058': email_AC, 'P-22/059': email_AC, 'P-22/060': email_AC,
                 'P-22/061': email_LB, 'P-22/062': email_SS, 'P-22/063': email_SS, 'P-22/064': email_LB,
                 'P-22/065': email_AC, 'P-22/066': email_AC, 'P-22/067': email_AC, 'P-22/068': email_AC,
                 'P-22/069': email_AC, 'P-22/070': email_SS, 'P-22/071': email_AC, 'P-22/072': email_LB,
                 'P-22/073': email_AC, 'P-22/074': email_LB, 'P-22/075': email_SS, 'P-22/076': email_LB,
                 'P-22/077': email_AC, 'P-22/078': email_AC, 'P-22/079': email_AC, 'P-22/080': email_SS,
                 'P-22/081': email_AC, 'P-22/082': email_LB, 'P-22/083': email_AC, 'P-22/084': email_LB,
                 'P-22/085': email_LB, 'P-22/086': email_LB, 'P-22/087': email_LB, 'P-22/088': email_LB,
                 'P-22/089': email_LB, 'P-22/090': email_LB, 'P-22/091': email_LB, 'P-22/092': email_LB,
                 'P-22/093': email_LB, 'P-22/094': email_LB, 'P-22/095': email_LB, 'P-22/096': email_LB,
                 'P-22/097': email_LB, 'P-22/098': email_LB, 'P-22/099': email_LB, 'P-22/100': email_LB,
                 'P-22/101': email_LB, 'P-22/102': email_LB, 'P-22/103': email_LB, 'P-22/104': email_LB,
                 'P-22/105': email_LB,
                 'P-23/001': email_LB, 'P-23/002': email_LB, 'P-23/003': email_LB, 'P-23/004': email_AC,
                 'P-23/005': email_AC, 'P-23/006': email_AC, 'P-23/007': email_LB, 'P-23/008': email_AC,
                 'P-23/009': email_AC, 'P-23/010': email_AC, 'P-23/011': email_SS, 'P-23/012': email_AC,
                 'P-23/013': email_LB, 'P-23/014': email_SS, 'P-23/015': email_AC, 'P-23/016': email_AC,
                 'P-23/017': email_SS, 'P-23/018': email_AC, 'P-23/019': email_LB, 'P-23/020': email_AC,
                 'P-23/021': email_LB, 'P-23/022': email_LB, 'P-23/023': email_AC, 'P-23/024': email_LB,
                 'P-23/025': email_LB, 'P-23/026': email_SS, 'P-23/027': email_LB, 'P-23/028': email_LB,
                 'P-23/029': email_LB, 'P-23/030': email_LB, 'P-23/031': email_AC, 'P-23/032': email_AC,
                 'P-23/033': email_AC, 'P-23/034': email_SS, 'P-23/035': email_AC, 'P-23/036': email_AC,
                 'P-23/037': email_LB, 'P-23/038': email_LB, 'P-23/039': email_LB, 'P-23/040': email_AC,
                 'P-23/041': email_AC, 'P-23/042': email_LB, 'P-23/043': email_LB, 'P-23/044': email_LB,
                 'P-23/045': email_AC, 'P-23/046': email_SS, 'P-23/047': email_AC, 'P-23/048': email_SS,
                 'P-23/049': email_LB, 'P-23/050': email_LB, 'P-23/051': email_AC, 'P-23/052': email_AC,
                 'P-23/053': email_AC, 'P-23/054': email_AC, 'P-23/055': email_AC, 'P-23/056': email_SS,
                 'P-23/057': email_LB, 'P-23/058': email_AC, 'P-23/059': email_LB, 'P-23/060': email_AC,
                 'P-23/061': email_LB, 'P-23/062': email_AC, 'P-23/063': email_AC, 'P-23/064': email_AC,
                 'P-23/065': email_AC, 'P-23/066': email_AC, 'P-23/067': email_AC, 'P-23/068': email_AC,
                 'P-23/069': email_AC, 'P-23/070': email_AC, 'P-23/071': email_AC, 'P-23/072': email_LB,
                 'P-23/073': email_AC, 'P-23/074': email_SS, 'P-23/075': email_LB, 'P-23/076': email_LB,
                 'P-23/077': email_AC, 'P-23/078': email_AC, 'P-23/079': email_LB, 'P-23/080': email_AC,
                 'P-23/081': email_AC, 'P-23/082': email_AC, 'P-23/083': email_AC, 'P-23/084': email_AC,
                 'P-23/085': email_AC, 'P-23/086': email_AC, 'P-23/087': email_AC, 'P-23/088': email_AC,
                 'P-23/089': email_SS, 'P-23/090': email_AC, 'P-23/091': email_AC, 'P-23/092': email_LB,
                 'P-23/093': email_AC, 'P-23/094': email_LB, 'P-23/095': email_AC, 'P-23/096': email_AC,
                 'P-23/097': email_AC, 'P-23/098': email_LB, 'P-23/099': email_LB, 'P-23/100': email_AC,
                 'P-23/101': email_AC, 'P-23/102': email_AC, 'P-23/103': email_LB, 'P-23/104': email_AC,
                 'P-23/105': email_SS, 'P-24/001': email_LB, 'P-24/002': email_LB, 'P-24/003': email_LB,
                 'P-24/004': email_AC, 'P-24/005': email_AC, 'P-24/006': email_AC, 'P-24/007': email_AC,
                 'P-24/008': email_AC, 'P-24/009': email_AC, 'P-24/010': email_AC, 'P-24/011': email_AC,
                 'P-24/012': email_SS, 'P-24/013': email_AC, 'P-24/014': email_AC, 'P-24/015': email_SS,
                 'P-24/016': email_AC, 'P-24/017': email_AC, 'P-24/018': email_AC, 'P-24/019': email_AC,
                 'P-24/020': email_AC, 'P-24/021': email_AC, 'P-24/022': email_AC, 'P-24/023': email_AC,
                 'P-24/024': email_AC, 'P-24/025': email_AC, 'P-24/026': email_AC, 'P-24/027': email_AC,
                 'P-24/028': email_AC, 'P-24/029': email_AC, 'P-24/030': email_AC, 'P-24/031': email_AC,
                 'P-24/032': email_AC, 'P-24/033': email_AC, 'P-24/034': email_AC, 'P-24/035': email_AC,
                 'P-24/036': email_AC, 'P-24/037': email_AC, 'P-24/038': email_AC, 'P-24/039': email_AC,
                 'P-24/040': email_AC, 'P-24/041': email_AC, 'P-24/042': email_AC, 'P-24/043': email_AC,
                 'P-24/044': email_AC, 'P-24/045': email_AC, 'P-24/046': email_AC, 'P-24/047': email_AC,
                 'P-24/048': email_AC, 'P-24/049': email_AC, 'P-24/050': email_AC, 'P-24/051': email_AC,
                   'P-24/052': email_AC, 'P-24/053': email_AC, 'P-24/054': email_AC, 'P-24/055': email_AC,
                   'P-24/056': email_AC, 'P-24/057': email_AC, 'P-24/058': email_AC, 'P-24/059': email_AC,
                   'P-24/060': email_AC, 'P-24/061': email_AC, 'P-24/062': email_AC, 'P-24/063': email_AC,
                   'P-24/064': email_AC, 'P-24/065': email_AC, 'P-24/066': email_LB, 'P-24/067': email_AC,
                   'P-24/068': email_AC, 'P-24/069': email_LB, 'P-24/070': email_LB, 'P-24/071': email_AC,
                   'P-24/072': email_AC, 'P-24/073': email_AC, 'P-24/074': email_AC, 'P-24/075': email_AC,
                   'P-24/076': email_AC, 'P-24/077': email_AC, 'P-24/078': email_AC, 'P-24/079': email_SS,
                   'P-24/080': email_SS, 'P-24/081': email_AC, 'P-24/082': email_AC, 'P-24/083': email_AC,
                   'P-24/084': email_AC, 'P-24/085': email_LB, 'P-24/086': email_CC, 'P-24/087': email_AC,
                   'P-24/088': email_AC, 'P-24/089': email_AC, 'P-24/090': email_AC, 'P-24/091': email_AC,
                   'P-24/092': email_SS, 'P-24/093': email_LB, 'P-24/094': email_LB, 'P-24/095': email_AC,
                   'P-24/096': email_CC, 'P-24/097': email_AC, 'P-24/098': email_CC, 'P-24/099': email_CC,
                   'P-24/100': email_SS, 'P-25/001': email_AC, 'P-25/002': email_AC, 'P-25/003': email_SS,
                   'P-25/004': email_AC, 'P-25/005': email_SS, 'P-25/006': email_CC, 'P-25/007': email_SS,
                   'P-25/008': email_AC, 'P-25/009': email_AC, 'P-25/010': email_AC, 'P-25/011': email_AC,
                   'P-25/012': email_SS, 'P-25/013': email_AC, 'P-25/014': email_AC, 'P-25/015': email_SS,
                   'P-25/016': email_AC, 'P-25/017': email_AC, 'P-25/018': email_AC, 'P-25/019': email_AC,
                   'P-25/020': email_AC, 'P-25/021': email_AC, 'P-25/022': email_AC, 'P-25/023': email_AC,
                   'P-25/024': email_AC, 'P-25/025': email_AC, 'P-25/026': email_SS, 'P-25/027': email_AC,
                   'P-25/028': email_AC, 'P-25/029': email_AC, 'P-25/030': email_AC, 'P-25/031': email_AC,
                   'P-25/032': email_AC, 'P-25/033': email_AC, 'P-25/034': email_AC, 'P-25/035': email_AC, }


def get_responsable_email(numero_pedido):
    """
            Función para identificar al responsable del pedido

            Args:
                df (pandas.DataFrame): DataFrame que contiene ['Nº pedido'] volcamos la columna a ['Responsable_email'] y transformamos con mapping

            Returns:
                pandas.DataFrame: DataFrame con columna ['Responsable_email'] en la que se encuentra el email del responsable del pedido
    """

    for key in email_mapping:
        if key in numero_pedido:
            return email_mapping[key]
    return None


def aplicar_estilos_y_guardar_excel(df, filename):
    # Crear un nuevo libro de trabajo y una hoja
    wb = Workbook()
    ws = wb.active

    # Definir los estilos
    cell_filling_blue_light = PatternFill(start_color="D4DCF4", end_color="D4DCF4", fill_type="solid")
    cell_filling = PatternFill(start_color="6678AF", end_color="6678AF", fill_type="solid")
    medium_dashed = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))
    font_white = Font(color='FFFFFF', bold=True)
    font_black = Font(color='000000')

    # Aplicar borde a todas las celdas
    for row in ws.iter_rows():
        for cell in row:
            cell.border = medium_dashed

    # Convertir el DataFrame a filas de la hoja de Excel
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        ws.append(row)
        for c_idx, cell in enumerate(row, 1):
            if r_idx == 0:
                # Aplicar estilo a la cabecera
                cell_obj = ws.cell(row=r_idx+1, column=c_idx)
                cell_obj.fill = cell_filling
                cell_obj.font = font_white
                cell_obj.border = medium_dashed
            else:
                # Aplicar estilo a las celdas de datos
                cell_obj = ws.cell(row=r_idx+1, column=c_idx)
                cell_obj.border = medium_dashed
                cell_obj.font = font_black
                cell_obj.fill = cell_filling_blue_light
                if isinstance(cell, pd.Timestamp):
                    cell_obj.fill = cell_filling_blue_light

        # Ajustar ancho de columna al contenido
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Aplicar autofiltro para la primera fila
    ws.auto_filter.ref = ws.dimensions

    # Guardar el archivo Excel
    wb.save(filename)


def aplicar_estilos_html(df):
    styles = {
        'fecha': 'background-color: #D4DCF4; text-align: left; font-size: 14px;',
        'header': 'background-color: #6678AF; color: #FFFFFF; text-align: left; font-size: 14px;',
        'cell_even': 'background-color: #D4DCF4; text-align: left; font-size: 14px;',
        'cell_default': 'background-color: #D4DCF4; text-align: left; font-size: 14px;'
    }

    def style_specific_cell(val):
        if isinstance(val, pd.Timestamp):
            return styles['fecha']
        return styles['cell_even']

    def apply_conditional_styles(val):
        if val == 'Rechazado':
            return 'color: #000000; font-weight: bold; background-color: #FFA19A; font-size: 14px;'
        elif val == 'Com. Menores':
            return 'color: #000000; font-weight: bold; background-color: #FFE5AD; font-size: 14px;'
        elif val == 'Com. Mayores':
            return 'color: #000000; font-weight: bold; background-color: #DBB054; font-size: 14px;'
        elif val == 'Comentado':
            return 'color: #000000; font-weight: bold; background-color: #F79646; font-size: 14px;'
        elif val == 'Aprobado':
            return 'color: #000000; font-weight: bold; background-color: #00D25F; font-size: 14px;'
        elif val == 'Eliminado':
            return 'color: #000000; font-weight: bold; background-color: #FF0000; font-size: 14px;'
        else:
            return 'text-align: left; font-size: 14px;'

    header_style = [{'selector': 'th', 'props': [('background-color', '#6678AF'),
                                                 ('color', '#FFFFFF'),
                                                 ('text-align', 'center'),
                                                 ('font-size', '14px'),
                                                 ('font-weight', 'bold')]}]

    # Aplicar los estilos con .map en lugar de .applymap
    styled = df.style \
        .map(style_specific_cell) \
        .map(apply_conditional_styles) \
        .set_table_styles(header_style)

    # Convertir a HTML sin índice
    return styled.to_html(index=False)


def aplicar_estilo_info(df):
    # Aplicar estilos básicos a todas las celdas
    estilo_celdas = 'background-color: #D4DCF4; text-align: left; font-size: 14px;'
    estilo_header = [{'selector': 'th', 'props': [('background-color', '#6678AF'),
                                                  ('color', '#FFFFFF'),
                                                  ('text-align', 'center'),
                                                  ('font-size', '15px'),
                                                  ('font-weight', 'bold')]}]

    styled = df.style \
        .map(lambda _: estilo_celdas) \
        .set_table_styles(estilo_header)

    return styled.to_html(index=False)